import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def export_to_excel(json_path, output_path):
    if not os.path.exists(json_path):
        print(f"[!] File not found: {json_path}")
        return

    # Read multi-line JSON
    data = []
    with open(json_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                try:
                    data.append(json.loads(line))
                except:
                    continue

    df = pd.DataFrame(data)
    if df.empty:
        print("[!] No data to export.")
        return

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Group by node (IP)
        for node_ip, group in df.groupby('node'):
            # Select and rename columns for a professional look
            export_df = group[[
                'check_id', 'title', 'status', 'severity', 
                'current_value', 'expected_value', 'remediation'
            ]].copy()
            
            # Xử lý: Nếu PASS thì không hiện hướng dẫn khắc phục
            export_df.loc[export_df['status'] == 'PASS', 'remediation'] = ""
            
            export_df.columns = [
                'ID', 'Hạng mục kiểm tra', 'Trạng thái', 'Mức độ', 
                'Giá trị thực tế', 'Giá trị kỳ vọng', 'Hướng dẫn khắc phục'
            ]
            
            # Write to a specific sheet
            sheet_name = f"Node_{node_ip.replace('.', '_')}"
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook and sheet for styling
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # --- STYLING ---
            # 1. Header Styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # 2. Body Styling & Conditional Coloring
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                status_cell = row[2] # Status column
                if status_cell.value == "PASS":
                    status_cell.font = Font(color="008000", bold=True) # Green
                elif status_cell.value == "FAIL":
                    status_cell.font = Font(color="FF0000", bold=True) # Red
                elif status_cell.value == "MANUAL":
                    status_cell.font = Font(color="0000FF", bold=True) # Blue
                
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 3. Column Widths
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 12
            worksheet.column_dimensions['E'].width = 30
            worksheet.column_dimensions['F'].width = 30
            worksheet.column_dimensions['G'].width = 50

    print(f"[OK] Báo cáo Excel đã được xuất thành công: {output_path}")

if __name__ == "__main__":
    JSON_FILE = "scripts/reports/cluster_results.json"
    OUTPUT_FILE = "CIS_Cassandra_Compliance_Report.xlsx"
    export_to_excel(JSON_FILE, OUTPUT_FILE)
